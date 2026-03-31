#!/usr/bin/env python3
"""
PM Workflow — Corporate XLSX Report Generator
Generates a professionally formatted project status report.
Runs locally with zero LLM token cost.

Usage:
  python3 .pm/scripts/generate-report.py
  python3 .pm/scripts/generate-report.py --output custom-name.xlsx

Requires: pip install openpyxl
Fallback: generates CSV files if openpyxl is not installed
"""

import json, re, os, sys, glob, csv
from datetime import datetime, date

OUTPUT_DIR = "specs/reports"

# ============================================================
# Data Readers
# ============================================================

def read_file(path):
    try:
        with open(path, "r") as f: return f.read()
    except FileNotFoundError: return ""

def read_json(path):
    try:
        with open(path, "r") as f: return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError): return {}

def parse_frontmatter(content):
    if not content.startswith("---"): return {}
    parts = content.split("---", 2)
    if len(parts) < 3: return {}
    fm = {}
    for line in parts[1].strip().split("\n"):
        if ":" in line:
            key, _, val = line.partition(":")
            key, val = key.strip(), val.strip().strip('"').strip("'")
            if val.startswith("[") and val.endswith("]"):
                val = [v.strip().strip('"').strip("'") for v in val[1:-1].split(",") if v.strip()]
            fm[key] = val
    return fm

def parse_requirements(content):
    reqs = []
    current = None
    for line in content.split("\n"):
        m = re.match(r"^## (REQ-\d+):\s*(.*)", line)
        if m:
            if current: reqs.append(current)
            current = {"id": m.group(1), "title": m.group(2), "lines": []}
        elif current:
            current["lines"].append(line)
    if current: reqs.append(current)
    return reqs

def parse_audit_log(path):
    entries = []
    for line in read_file(path).strip().split("\n"):
        if line.strip():
            try: entries.append(json.loads(line))
            except json.JSONDecodeError: pass
    return entries

def get_files(pattern, key_fn=None):
    results = []
    for f in sorted(glob.glob(pattern)):
        content = read_file(f)
        fm = parse_frontmatter(content)
        fm["_file"] = os.path.basename(f)
        fm["_path"] = f
        fm["_content"] = content
        results.append(fm)
    return results

def get_signoff_docs():
    docs = []
    for name, path in [("SRS", "specs/srs/srs.md"), ("System Design", "specs/design/system-design.md"),
                        ("Sequence Diagrams", "specs/design/sequence-diagrams.md"), ("Test Plan", "specs/test-plan/test-plan.md")]:
        if os.path.exists(path):
            fm = parse_frontmatter(read_file(path))
            docs.append({"name": name, "path": path, "status": fm.get("status", "draft"),
                         "approved_by": fm.get("approved_by", ""), "approved_date": fm.get("approved_date", ""),
                         "created": fm.get("created", ""), "updated": fm.get("updated", "")})
    for f in sorted(glob.glob("specs/journeys/journey-*.md")):
        fm = parse_frontmatter(read_file(f))
        docs.append({"name": f"User Journey: {os.path.basename(f).replace('journey-','').replace('.md','')}", "path": f,
                     "status": fm.get("status", "draft"), "approved_by": fm.get("approved_by", ""),
                     "approved_date": fm.get("approved_date", ""), "created": fm.get("created", ""), "updated": fm.get("updated", "")})
    return docs

def parse_deliverable_tracker():
    items = []
    in_table = header_seen = False
    for line in read_file("specs/deliverable-tracker.md").split("\n"):
        if "| ID |" in line or "| id |" in line.lower(): in_table, header_seen = True, False; continue
        if in_table and "|---" in line: header_seen = True; continue
        if in_table and header_seen and line.startswith("|"):
            cols = [c.strip() for c in line.split("|")[1:-1]]
            if len(cols) >= 7:
                items.append({"id": cols[0], "name": cols[1], "role": cols[2], "owner": cols[3],
                             "reqs": cols[4], "due": cols[5], "status": cols[6]})
        elif in_table and header_seen and not line.startswith("|"): in_table = False
    return items

def parse_ingestion_log():
    entries, current = [], None
    for line in read_file(".pm/ingestion-log.md").split("\n"):
        if line.startswith("## Ingestion:"):
            if current: entries.append(current)
            current = {"date": line.replace("## Ingestion:", "").strip(), "lines": []}
        elif current: current["lines"].append(line)
    if current: entries.append(current)
    return entries


# ============================================================
# XLSX Corporate Template
# ============================================================

def generate_xlsx(output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # -- Design tokens --
    NAVY    = "1B2A4A"
    NAVY_LT = "2C3E6B"
    STEEL   = "4A5568"
    SLATE   = "2D3748"
    BORDER  = "CBD5E0"
    BG_ALT  = "F7FAFC"
    BG_SEC  = "EDF2F7"
    BG_MET  = "EBF8FF"
    WHITE   = "FFFFFF"

    STATUS_MAP = {
        "done":     ("C6F6D5", "22543D"), "complete": ("C6F6D5", "22543D"), "completed": ("C6F6D5", "22543D"),
        "closed":   ("C6F6D5", "22543D"), "approved": ("C6F6D5", "22543D"),
        "active":   ("BEE3F8", "2A4365"), "in-progress": ("BEE3F8", "2A4365"), "in_progress": ("BEE3F8", "2A4365"),
        "in progress": ("BEE3F8", "2A4365"),
        "in review": ("E9D8FD", "44337A"), "review": ("E9D8FD", "44337A"),
        "blocked":  ("FED7D7", "742A2A"), "needs-decision": ("FED7D7", "742A2A"),
        "draft":    ("FEFCBF", "744210"), "backlog": ("FEFCBF", "744210"), "open": ("FEFCBF", "744210"),
        "not started": ("FEFCBF", "744210"),
    }

    # -- Reusable styles --
    bdr = Border(left=Side("thin", BORDER), right=Side("thin", BORDER), top=Side("thin", BORDER), bottom=Side("thin", BORDER))
    bdr_none = Border()

    def F(bold=False, color=SLATE, size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def BG(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def A(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def status_style(val):
        s = str(val).lower().strip()
        bg, fg = STATUS_MAP.get(s, (WHITE, SLATE))
        return BG(bg), F(bold=True, color=fg, size=9)

    # -- Sheet builder --
    def make_sheet(ws, title, headers, rows, widths, status_col=None):
        """Build a complete corporate-styled sheet."""
        ncols = len(headers)

        # --- Title bar (row 1) ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=f"  {title}")
        c.font = F(bold=True, color=WHITE, size=13)
        c.fill = BG(NAVY)
        c.alignment = A("left", "center")
        ws.row_dimensions[1].height = 36
        for ci in range(2, ncols + 1):
            ws.cell(row=1, column=ci).fill = BG(NAVY)

        # --- Subtitle bar (row 2) ---
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        state = read_json(".pm/state.json")
        pname = state.get("project_name", "Project")
        c2 = ws.cell(row=2, column=1, value=f"  {pname}  |  {datetime.now().strftime('%B %d, %Y')}")
        c2.font = F(color=WHITE, size=9)
        c2.fill = BG(NAVY_LT)
        c2.alignment = A("left", "center")
        ws.row_dimensions[2].height = 22
        for ci in range(2, ncols + 1):
            ws.cell(row=2, column=ci).fill = BG(NAVY_LT)

        # --- Spacer (row 3) ---
        ws.row_dimensions[3].height = 6

        # --- Column headers (row 4) ---
        hr = 4
        ws.row_dimensions[hr].height = 28
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=hr, column=ci, value=h.upper())
            c.font = F(bold=True, color=WHITE, size=9)
            c.fill = BG(NAVY)
            c.alignment = A("center", "center", wrap=True)
            c.border = bdr

        # --- Data rows (row 5+) ---
        for ri, row in enumerate(rows):
            actual = hr + 1 + ri
            ws.row_dimensions[actual].height = 24
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=actual, column=ci, value=val)
                c.font = F(size=10)
                c.alignment = A("left", "center", wrap=True)
                c.border = bdr
                if status_col and ci == status_col:
                    sf, ff = status_style(val)
                    c.fill = sf
                    c.font = ff
                    c.alignment = A("center", "center")
                else:
                    c.fill = BG(BG_ALT) if ri % 2 == 0 else BG(WHITE)

        # --- Column widths ---
        from openpyxl.utils import get_column_letter
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # --- Freeze & print ---
        ws.freeze_panes = f"A{hr + 1}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    # -- Dashboard builder --
    def make_dashboard(ws, state, reqs, tasks, personas, signoff, deliverables, active_epic):
        ncols = 6
        pname = state.get("project_name", "Project")
        phase = state.get("phase", 0)
        phase_name = state.get("phase_name", "Setup")
        total_reqs = len(reqs)
        total_tasks = len(tasks)
        done_tasks = sum(1 for t in tasks if t.get("status", "").lower() in ("closed", "completed", "done"))
        active_tasks = sum(1 for t in tasks if t.get("status", "").lower() in ("in-progress", "in_progress"))
        blocked_count = len(state.get("blocked", []))
        signoff_done = sum(1 for d in signoff if d.get("status", "").lower() == "approved")
        deliv_done = sum(1 for d in deliverables if d.get("status", "").lower() == "approved")
        pct = round(done_tasks / total_tasks * 100) if total_tasks else 0

        from openpyxl.utils import get_column_letter
        for ci in range(1, 7):
            ws.column_dimensions[get_column_letter(ci)].width = 22

        r = 1  # current row

        # === TITLE BAR ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=f"  {pname}")
        c.font = F(bold=True, color=WHITE, size=18)
        c.fill = BG(NAVY)
        c.alignment = A("left", "center")
        ws.row_dimensions[r].height = 50
        for ci in range(2, ncols + 1): ws.cell(row=r, column=ci).fill = BG(NAVY)

        # === SUBTITLE ===
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=f"  Project Status Report  |  {datetime.now().strftime('%B %d, %Y')}  |  Phase {phase}: {phase_name}")
        c.font = F(color=WHITE, size=10)
        c.fill = BG(NAVY_LT)
        c.alignment = A("left", "center")
        ws.row_dimensions[r].height = 28
        for ci in range(2, ncols + 1): ws.cell(row=r, column=ci).fill = BG(NAVY_LT)

        # === SPACER ===
        r += 1
        ws.row_dimensions[r].height = 12

        # === SECTION: KEY METRICS ===
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value="  KEY METRICS")
        c.font = F(bold=True, color=NAVY, size=11)
        c.fill = BG(BG_SEC)
        c.alignment = A("left", "center")
        ws.row_dimensions[r].height = 26
        for ci in range(2, ncols + 1): ws.cell(row=r, column=ci).fill = BG(BG_SEC)

        # Metric cards: value row
        r += 1
        ws.row_dimensions[r].height = 40
        metrics = [(f"{total_reqs}", "Requirements"), (f"{total_tasks}", "Total Tasks"), (f"{done_tasks}", "Completed"),
                   (f"{active_tasks}", "In Progress"), (f"{blocked_count}", "Blocked"), (f"{pct}%", "Completion")]
        for ci, (val, _) in enumerate(metrics, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = F(bold=True, color=NAVY, size=18)
            c.fill = BG(BG_MET)
            c.alignment = A("center", "center")
            c.border = Border(left=Side("thin", "E2E8F0"), right=Side("thin", "E2E8F0"), top=Side("thin", "E2E8F0"))

        # Metric cards: label row
        r += 1
        ws.row_dimensions[r].height = 20
        for ci, (_, label) in enumerate(metrics, 1):
            c = ws.cell(row=r, column=ci, value=label)
            c.font = F(color=STEEL, size=9)
            c.fill = BG(BG_MET)
            c.alignment = A("center", "center")
            c.border = Border(left=Side("thin", "E2E8F0"), right=Side("thin", "E2E8F0"), bottom=Side("thin", "E2E8F0"))

        # === SPACER ===
        r += 1
        ws.row_dimensions[r].height = 12

        # === SECTION: PHASE PROGRESS ===
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value="  PHASE PROGRESS")
        c.font = F(bold=True, color=NAVY, size=11)
        c.fill = BG(BG_SEC)
        c.alignment = A("left", "center")
        ws.row_dimensions[r].height = 26
        for ci in range(2, ncols + 1): ws.cell(row=r, column=ci).fill = BG(BG_SEC)

        # Phase names
        r += 1
        ws.row_dimensions[r].height = 24
        phases = ["Ingest", "Brainstorm", "Document", "Plan", "Execute", "Track"]
        for ci, p in enumerate(phases, 1):
            c = ws.cell(row=r, column=ci, value=p)
            c.font = F(bold=True, size=10)
            c.alignment = A("center", "center")
            c.border = bdr

        # Phase statuses
        r += 1
        ws.row_dimensions[r].height = 24
        for ci, p in enumerate(phases, 1):
            idx = ci - 1
            txt = "Complete" if idx < phase else ("Current" if idx == phase else "Upcoming")
            c = ws.cell(row=r, column=ci, value=txt)
            sf, ff = status_style("done" if idx < phase else ("active" if idx == phase else "draft"))
            c.fill = sf
            c.font = ff
            c.alignment = A("center", "center")
            c.border = bdr

        # === SPACER ===
        r += 1
        ws.row_dimensions[r].height = 12

        # === SECTION: PROJECT DETAILS ===
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value="  PROJECT DETAILS")
        c.font = F(bold=True, color=NAVY, size=11)
        c.fill = BG(BG_SEC)
        c.alignment = A("left", "center")
        ws.row_dimensions[r].height = 26
        for ci in range(2, ncols + 1): ws.cell(row=r, column=ci).fill = BG(BG_SEC)

        # Detail pairs (label, value) x 3 per row
        detail_rows = [
            ("Active Epic", active_epic or "None", "PRD Status", "Created" if os.path.exists("specs/prd/prd.md") else "Not created", "Personas", f"{len(personas)} defined"),
            ("Sign-Off Docs", f"{signoff_done}/{len(signoff)} approved", "Deliverables", f"{deliv_done}/{len(deliverables)} approved", "Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for detail_row in detail_rows:
            r += 1
            ws.row_dimensions[r].height = 24
            for ci in range(0, 6, 2):
                lc = ws.cell(row=r, column=ci + 1, value=detail_row[ci])
                lc.font = F(bold=True, color=STEEL, size=9)
                lc.fill = BG(BG_ALT)
                lc.alignment = A("left", "center")
                lc.border = bdr
                vc = ws.cell(row=r, column=ci + 2, value=detail_row[ci + 1])
                vc.font = F(size=10)
                vc.fill = BG(WHITE)
                vc.alignment = A("left", "center")
                vc.border = bdr

        # Print
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ============================================================
    # Load all data
    # ============================================================
    state = read_json(".pm/state.json")
    context = read_file(".pm/context.md")
    reqs = parse_requirements(read_file("specs/requirements.md"))
    audit = parse_audit_log(".pm/audit.log")
    signoff = get_signoff_docs()
    deliverables = parse_deliverable_tracker()
    personas = get_files("specs/personas/*.md")
    stories = get_files("specs/stories/us-*.md")
    meetings = get_files(".pm/meeting-prep-*.md")
    ingestions = parse_ingestion_log()

    active_epic = state.get("active_epic", "")
    epic_dir = f"specs/epics/{active_epic}" if active_epic else ""
    tasks = get_files(f"{epic_dir}/[0-9]*.md") if epic_dir and os.path.isdir(epic_dir) else []
    if not tasks:
        for d in sorted(glob.glob("specs/epics/*/")):
            t = get_files(f"{d}[0-9]*.md")
            if t: tasks, epic_dir, active_epic = t, d.rstrip("/"), os.path.basename(d.rstrip("/")); break

    # ============================================================
    # Build Sheets
    # ============================================================

    # 1. Dashboard
    ws1 = wb.active
    ws1.title = "Dashboard"
    make_dashboard(ws1, state, reqs, tasks, personas, signoff, deliverables, active_epic)

    # 2. Requirements
    ws2 = wb.create_sheet("Requirements")
    req_rows = []
    for r in reqs:
        priority = status = source = ""
        for line in r.get("lines", []):
            if "priority" in line.lower() and ":" in line: priority = line.split(":", 1)[1].strip()
            if "needs-decision" in line.lower(): status = "needs-decision"
            if "source" in line.lower() and ":" in line: source = line.split(":", 1)[1].strip()
        req_rows.append((r["id"], r["title"], status or "active", priority, source))
    make_sheet(ws2, "Requirements", ["ID", "Title", "Status", "Priority", "Source"],
               req_rows or [("—", "No requirements", "", "", "")], [12, 45, 14, 12, 35], status_col=3)

    # 3. PRD Summary
    ws3 = wb.create_sheet("PRD Summary")
    prd_rows = []
    prd_content = read_file("specs/prd/prd.md")
    for line in prd_content.split("\n"):
        m = re.match(r"^### (?:Feature \d+:\s*|[\d]+\.\s*)(.*)", line)
        if m: prd_rows.append((m.group(1), "", "", ""))
    make_sheet(ws3, "PRD Summary", ["Feature", "Status", "Priority", "REQ IDs"],
               prd_rows or [("—", "No PRD", "", "")], [45, 14, 12, 25])

    # 4. Personas
    ws4 = wb.create_sheet("Personas")
    p_rows = [(p.get("name", p["_file"]), p.get("type", ""), p.get("requirement", ""), p.get("created", "")) for p in personas]
    make_sheet(ws4, "Personas", ["Name", "Type", "Linked REQ", "Created"],
               p_rows or [("—", "No personas", "", "")], [28, 14, 14, 20])

    # 5. Discovery & Strategy
    ws5 = wb.create_sheet("Discovery & Strategy")
    s_rows = []
    in_d = in_q = False
    for line in context.split("\n"):
        if "## Key Decisions" in line: in_d, in_q = True, False; continue
        if "## Open Questions" in line: in_q, in_d = True, False; continue
        if line.startswith("## "): in_d = in_q = False; continue
        if in_d and line.strip().startswith("- "): s_rows.append(("Decision", line.strip("- ").strip()))
        if in_q and line.strip().startswith("- "): s_rows.append(("Open Question", line.strip("- ").strip()))
    for f in ["strategy/positioning.md", "strategy/roadmap.md"]:
        if os.path.exists(f): s_rows.append(("Artifact", f))
    make_sheet(ws5, "Discovery & Strategy", ["Type", "Details"],
               s_rows or [("—", "No data")], [18, 65])

    # 6. User Stories
    ws6 = wb.create_sheet("User Stories")
    st_rows = [(s["_file"].replace(".md", ""), s.get("name", ""), s.get("status", "open"), s.get("epic", "")) for s in stories]
    make_sheet(ws6, "User Stories", ["ID", "Title", "Status", "Epic"],
               st_rows or [("—", "No stories", "", "")], [14, 40, 14, 35], status_col=3)

    # 7. Epic & Tasks
    ws7 = wb.create_sheet("Epic & Tasks")
    t_rows = []
    for t in tasks:
        dep = t.get("depends_on", "")
        if isinstance(dep, list): dep = ", ".join(str(d) for d in dep)
        eff = t.get("effort", "")
        if isinstance(eff, dict): eff = eff.get("size", "")
        t_rows.append((t["_file"].replace(".md", ""), t.get("name", ""), t.get("status", "open"), dep, eff, t.get("created", ""), t.get("updated", "")))
    make_sheet(ws7, "Epic & Tasks", ["ID", "Name", "Status", "Depends On", "Effort", "Created", "Updated"],
               t_rows or [("—", "No tasks", "", "", "", "", "")], [10, 35, 14, 14, 10, 18, 18], status_col=3)

    # 8. Timeline
    ws8 = wb.create_sheet("Timeline")
    tl_rows = []
    for t in tasks:
        days = t.get("effort", "")
        if isinstance(days, dict): days = days.get("days", "")
        tl_rows.append((t["_file"].replace(".md",""), t.get("name",""), t.get("status","open"), t.get("created",""), t.get("started", t.get("created","")), t.get("completed",""), days))
    make_sheet(ws8, "Timeline", ["ID", "Name", "Status", "Created", "Started", "Completed", "Est. Days"],
               tl_rows or [("—", "No data", "", "", "", "", "")], [10, 35, 14, 18, 18, 18, 10], status_col=3)

    # 9. Deliverables
    ws9 = wb.create_sheet("Deliverables")
    d_rows = [(d["id"], d["name"], d["role"], d["owner"], d["reqs"], d["due"], d["status"]) for d in deliverables]
    make_sheet(ws9, "Deliverables", ["ID", "Name", "Role", "Owner", "REQ IDs", "Due Date", "Status"],
               d_rows or [("—", "No deliverables", "", "", "", "", "")], [10, 30, 14, 14, 14, 14, 14], status_col=7)

    # 10. Sign-Off Status
    ws10 = wb.create_sheet("Sign-Off Status")
    so_rows = [(d["name"], d["status"], d["approved_by"], d["approved_date"], d["updated"]) for d in signoff]
    make_sheet(ws10, "Sign-Off Status", ["Document", "Status", "Approved By", "Approved Date", "Last Updated"],
               so_rows or [("—", "No docs", "", "", "")], [30, 14, 20, 18, 18], status_col=2)

    # 11. Risks & Blockers
    ws11 = wb.create_sheet("Risks & Blockers")
    b_rows = []
    for b in state.get("blocked", []):
        b_rows.append(("Manual Block", b.get("description", ""), b.get("since", ""), b.get("blocked_by", "")))
    for t in tasks:
        if t.get("status", "").lower() in ("closed", "completed", "in-progress", "in_progress"): continue
        deps = t.get("depends_on", [])
        if isinstance(deps, str): deps = [d.strip() for d in deps.strip("[]").split(",") if d.strip()]
        for dep in deps:
            df = os.path.join(epic_dir, f"{str(dep).zfill(3)}.md")
            if os.path.exists(df):
                dfm = parse_frontmatter(read_file(df))
                if dfm.get("status", "").lower() not in ("closed", "completed"):
                    b_rows.append(("Dependency", f"Task {t['_file']} waits on {dep}", "", f"{dep}: {dfm.get('status','unknown')}"))
    make_sheet(ws11, "Risks & Blockers", ["Type", "Description", "Since", "Blocked By"],
               b_rows or [("None", "No blockers", "", "")], [16, 45, 14, 25], status_col=1)

    # 12. Traceability
    ws12 = wb.create_sheet("Traceability")
    tr_rows = []
    epic_fm = parse_frontmatter(read_file(os.path.join(epic_dir, "epic.md"))) if epic_dir else {}
    epic_reqs = epic_fm.get("requirements", [])
    if isinstance(epic_reqs, str): epic_reqs = [epic_reqs]
    prd_path = epic_fm.get("prd", "specs/prd/prd.md")
    for t in tasks:
        tr_rows.append((", ".join(epic_reqs) if epic_reqs else "", prd_path, f"{active_epic}/epic.md", t["_file"].replace(".md",""), t.get("name",""), t.get("status","")))
    make_sheet(ws12, "Traceability Matrix", ["REQ IDs", "PRD", "Epic", "Task", "Task Name", "Status"],
               tr_rows or [("", "", "", "", "No data", "")], [14, 25, 22, 10, 30, 14], status_col=6)

    # 13. Ingestion Log
    ws13 = wb.create_sheet("Ingestion Log")
    ig_rows = []
    for e in ingestions:
        src = req_r = ""
        for line in e.get("lines", []):
            if "Source:" in line: src = line.split("Source:", 1)[1].strip()
            if "REQ IDs Created:" in line: req_r = line.split("REQ IDs Created:", 1)[1].strip()
        ig_rows.append((e.get("date", ""), src, req_r))
    make_sheet(ws13, "Ingestion Log", ["Date", "Source", "REQ IDs Created"],
               ig_rows or [("—", "No ingestions", "")], [18, 50, 25])

    # 14. Meeting Prep
    ws14 = wb.create_sheet("Meeting Prep")
    m_rows = [(m.get("name", m["_file"]), m.get("meeting_type", ""), m.get("company", ""), m.get("created", ""), m.get("attendees", "")) for m in meetings]
    make_sheet(ws14, "Meeting Prep", ["Topic", "Type", "Company", "Created", "Attendees"],
               m_rows or [("—", "No meetings", "", "", "")], [30, 14, 20, 18, 30])

    # 15. Activity Log
    ws15 = wb.create_sheet("Activity Log")
    a_rows = [(a.get("timestamp",""), f"Phase {a.get('phase','')}", a.get("action", a.get("skill","")),
               ", ".join(a.get("artifacts_created",[])), a.get("req_id", a.get("reason",""))) for a in audit[-50:]]
    make_sheet(ws15, "Activity Log", ["Timestamp", "Phase", "Action", "Artifacts", "Details"],
               a_rows or [("—", "", "No activity", "", "")], [22, 10, 25, 40, 25])

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================
# CSV Fallback
# ============================================================

def generate_csv_fallback():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    state = read_json(".pm/state.json")
    reqs = parse_requirements(read_file("specs/requirements.md"))
    active_epic = state.get("active_epic", "")
    tasks = get_files(f"specs/epics/{active_epic}/[0-9]*.md") if active_epic else []

    with open(os.path.join(OUTPUT_DIR, "requirements.csv"), "w", newline="") as f:
        w = csv.writer(f); w.writerow(["REQ ID", "Title"])
        for r in reqs: w.writerow([r["id"], r["title"]])

    with open(os.path.join(OUTPUT_DIR, "tasks.csv"), "w", newline="") as f:
        w = csv.writer(f); w.writerow(["Task ID", "Name", "Status", "Depends On"])
        for t in tasks: w.writerow([t["_file"], t.get("name",""), t.get("status",""), t.get("depends_on","")])

    print(f"CSV files saved to {OUTPUT_DIR}/")
    print("Install openpyxl for full XLSX: pip install openpyxl")


# ============================================================
# Main
# ============================================================

if __name__ == "__main__":
    state = read_json(".pm/state.json")
    pname = state.get("project_name", "Project").lower().replace(" ", "-")
    default_file = f"{pname}-status-{date.today().isoformat()}.xlsx"
    output = os.path.join(OUTPUT_DIR, default_file)

    if len(sys.argv) > 2 and sys.argv[1] == "--output":
        output = os.path.join(OUTPUT_DIR, sys.argv[2])

    try:
        import openpyxl
        result = generate_xlsx(output)
        print(f"✅ Report generated: {result}")
        print(f"   15 tabs | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print(f"\n   Open in Excel or upload to Google Sheets.")
    except ImportError:
        print("⚠️  openpyxl not installed. Generating CSV fallback...")
        generate_csv_fallback()
