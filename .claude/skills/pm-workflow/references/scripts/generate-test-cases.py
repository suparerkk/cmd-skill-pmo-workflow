#!/usr/bin/env python3
"""
PM Workflow — Deterministic Test Case Generator
Reads specs/behavior/REQ-*.md → generates XLSX with UAT, SIT, E2E tabs.
Same input = same output. No LLM involvement.

Usage:
  python3 .pm/scripts/generate-test-cases.py
  python3 .pm/scripts/generate-test-cases.py --output custom-name.xlsx

Requires: pip install openpyxl (falls back to CSV if not installed)
"""

import json, os, sys, csv, re, glob
from datetime import datetime, date


OUTPUT_DIR = "specs/test-cases"
PROJECT_DATA = ".pm/project-data.json"
STATE_FILE = ".pm/state.json"
BEHAVIOR_GLOB = "specs/behavior/REQ-*.md"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def read_json(path):
    try:
        with open(path, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def get_project_name():
    d = read_json(PROJECT_DATA)
    name = d.get("project_name", "")
    if name:
        return name
    d = read_json(STATE_FILE)
    return d.get("project_name", "Project")


def parse_frontmatter(text):
    """Extract YAML-like frontmatter between --- delimiters."""
    fm = {}
    m = re.match(r"^---\s*\n(.*?)\n---", text, re.DOTALL)
    if not m:
        return fm
    for line in m.group(1).splitlines():
        if ":" in line:
            key, _, val = line.partition(":")
            fm[key.strip()] = val.strip()
    return fm


def parse_scenarios_table(text):
    """Find the ## Scenarios section and parse its markdown table."""
    # Find the Scenarios heading
    m = re.search(r"^##\s+Scenarios\s*$", text, re.MULTILINE)
    if not m:
        return []
    rest = text[m.end():]

    # Find the header row (starts with | ID)
    lines = rest.splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("|") and re.search(r"\bID\b", stripped):
            header_idx = i
            break
    if header_idx is None:
        return []

    # Parse header columns
    header_line = lines[header_idx]
    headers = [c.strip() for c in header_line.strip().strip("|").split("|")]

    # Skip separator line (|---|---|...)
    data_start = header_idx + 2

    scenarios = []
    for line in lines[data_start:]:
        stripped = line.strip()
        if not stripped.startswith("|"):
            break  # end of table
        cols = [c.strip() for c in stripped.strip("|").split("|")]
        if len(cols) < len(headers):
            cols.extend([""] * (len(headers) - len(cols)))
        row = {}
        for h, v in zip(headers, cols):
            row[h.lower()] = v
        scenarios.append(row)

    return scenarios


def parse_behavior_specs():
    """Read all specs/behavior/REQ-*.md files sorted by filename."""
    files = sorted(glob.glob(BEHAVIOR_GLOB))
    all_scenarios = []

    for fpath in files:
        with open(fpath, "r") as f:
            text = f.read()

        fm = parse_frontmatter(text)
        req_id = fm.get("req_id", "")
        title = fm.get("title", "")

        scenarios = parse_scenarios_table(text)
        for s in scenarios:
            s["_req_id"] = req_id
            s["_title"] = title
            s["_file"] = fpath
        all_scenarios.extend(scenarios)

    return all_scenarios, len(files)


def route_scenarios(scenarios):
    """Split scenarios into UAT, SIT, E2E buckets based on Type column."""
    uat, sit, e2e = [], [], []
    for s in scenarios:
        types_raw = s.get("type", "")
        types = [t.strip().upper() for t in types_raw.split(",")]
        for t in types:
            if t == "UAT":
                uat.append(s)
            elif t == "SIT":
                sit.append(s)
            elif t == "E2E":
                e2e.append(s)
    return uat, sit, e2e


def make_tc_id(prefix, req_id, scenario_id):
    """Build TC ID like UAT-REQ001-S01."""
    clean_req = req_id.replace("-", "").replace("_", "")
    return f"{prefix}-{clean_req}-{scenario_id}"


def number_steps(text, delimiter):
    """Split text by delimiter and number each part."""
    parts = [p.strip() for p in text.split(delimiter) if p.strip()]
    if len(parts) <= 1:
        return text.strip()
    return "\n".join(f"{i}. {p}" for i, p in enumerate(parts, 1))


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------

def build_uat_row(s):
    req_id = s["_req_id"]
    sid = s.get("id", "")
    given = s.get("given", "")
    when = s.get("when", "")
    then = s.get("then", "")
    return [
        make_tc_id("UAT", req_id, sid),
        req_id,
        s["_title"],
        f"When {when}",
        given,
        number_steps(when, ","),
        then,
        "Not Run",
        "",
        "",
    ]


def build_sit_row(s):
    req_id = s["_req_id"]
    sid = s.get("id", "")
    given = s.get("given", "")
    when = s.get("when", "")
    then = s.get("then", "")
    return [
        make_tc_id("SIT", req_id, sid),
        req_id,
        s["_title"],
        s["_title"],  # Component: default to REQ title
        given,
        when,
        then,
        "",
        "Not Run",
        "",
        "",
    ]


def build_e2e_row(s):
    req_id = s["_req_id"]
    sid = s.get("id", "")
    given = s.get("given", "")
    when = s.get("when", "")
    then = s.get("then", "")
    # Split steps on arrow or comma
    if "\u2192" in when:
        steps = number_steps(when, "\u2192")
    else:
        steps = number_steps(when, ",")
    return [
        make_tc_id("E2E", req_id, sid),
        req_id,
        s["_title"],
        when,
        given,
        steps,
        then,
        "",
        "Not Run",
        "",
        "",
    ]


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

UAT_HEADERS = ["TC ID", "REQ ID", "REQ Title", "Scenario", "Precondition",
               "Steps", "Expected Result", "Status", "Tested By", "Notes"]
UAT_WIDTHS = [22, 12, 28, 30, 28, 30, 35, 12, 14, 20]

SIT_HEADERS = ["TC ID", "REQ ID", "REQ Title", "Component", "Precondition",
               "Input", "Expected Output", "Integration Points", "Status",
               "Tested By", "Notes"]
SIT_WIDTHS = [22, 12, 28, 22, 28, 30, 35, 22, 12, 14, 20]

E2E_HEADERS = ["TC ID", "REQ ID", "REQ Title", "Flow Name", "Preconditions",
               "Steps", "Expected Result", "Data Requirements", "Status",
               "Tested By", "Notes"]
E2E_WIDTHS = [22, 12, 28, 30, 28, 35, 35, 22, 12, 14, 20]


def generate_xlsx(output_path, uat_rows, sit_rows, e2e_rows, project_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Corporate colors — matches generate-report.py
    NAVY = "1B2A4A"
    NAVY_LT = "2C3E6B"
    SLATE = "2D3748"
    BORDER_C = "CBD5E0"
    BG_ALT = "F7FAFC"
    WHITE = "FFFFFF"
    YELLOW_BG = "FEFCBF"
    YELLOW_FG = "744210"

    bdr = Border(
        left=Side("thin", BORDER_C), right=Side("thin", BORDER_C),
        top=Side("thin", BORDER_C), bottom=Side("thin", BORDER_C),
    )

    def F(bold=False, color=SLATE, size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def BG(c):
        return PatternFill(start_color=c, end_color=c, fill_type="solid")

    def A(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def status_style(val):
        s = str(val).lower().strip()
        if s == "not run":
            return BG(YELLOW_BG), F(bold=True, color=YELLOW_FG, size=9)
        return BG(WHITE), F(size=10)

    def build_sheet(ws, tab_title, headers, rows, widths, status_col):
        nc = len(headers)

        # Row 1: Title bar
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
        c = ws.cell(row=1, column=1, value=f"  Test Cases \u2014 {tab_title}")
        c.font = F(True, WHITE, 13)
        c.fill = BG(NAVY)
        c.alignment = A("left")
        ws.row_dimensions[1].height = 36
        for ci in range(2, nc + 1):
            ws.cell(row=1, column=ci).fill = BG(NAVY)

        # Row 2: Subtitle
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
        c2 = ws.cell(row=2, column=1,
                      value=f"  {project_name}  |  {datetime.now().strftime('%B %d, %Y')}")
        c2.font = F(color=WHITE, size=9)
        c2.fill = BG(NAVY_LT)
        c2.alignment = A("left")
        ws.row_dimensions[2].height = 22
        for ci in range(2, nc + 1):
            ws.cell(row=2, column=ci).fill = BG(NAVY_LT)

        # Row 3: Spacer
        ws.row_dimensions[3].height = 6

        # Row 4: Header
        hr = 4
        ws.row_dimensions[hr].height = 28
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=hr, column=ci, value=h.upper())
            c.font = F(True, WHITE, 9)
            c.fill = BG(NAVY)
            c.alignment = A("center", "center", True)
            c.border = bdr

        # Data rows
        for ri, row in enumerate(rows):
            ar = hr + 1 + ri
            ws.row_dimensions[ar].height = 24
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ar, column=ci, value=val)
                c.font = F(size=10)
                c.alignment = A("left", "center", True)
                c.border = bdr
                if ci == status_col:
                    sf, ff = status_style(val)
                    c.fill = sf
                    c.font = ff
                    c.alignment = A("center", "center")
                else:
                    c.fill = BG(BG_ALT) if ri % 2 == 0 else BG(WHITE)

        # Column widths
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Freeze panes below header
        ws.freeze_panes = f"A{hr + 1}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # --- UAT tab (first sheet) ---
    ws_uat = wb.active
    ws_uat.title = "UAT"
    build_sheet(ws_uat, "UAT", UAT_HEADERS, uat_rows, UAT_WIDTHS, status_col=8)

    # --- SIT tab ---
    ws_sit = wb.create_sheet("SIT")
    build_sheet(ws_sit, "SIT", SIT_HEADERS, sit_rows, SIT_WIDTHS, status_col=9)

    # --- E2E tab ---
    ws_e2e = wb.create_sheet("E2E")
    build_sheet(ws_e2e, "E2E", E2E_HEADERS, e2e_rows, E2E_WIDTHS, status_col=9)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CSV fallback
# ---------------------------------------------------------------------------

def generate_csv_fallback(uat_rows, sit_rows, e2e_rows):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = date.today().isoformat()

    path_uat = os.path.join(OUTPUT_DIR, f"test-cases-uat-{today}.csv")
    with open(path_uat, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(UAT_HEADERS)
        w.writerows(uat_rows)

    path_sit = os.path.join(OUTPUT_DIR, f"test-cases-sit-{today}.csv")
    with open(path_sit, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(SIT_HEADERS)
        w.writerows(sit_rows)

    path_e2e = os.path.join(OUTPUT_DIR, f"test-cases-e2e-{today}.csv")
    with open(path_e2e, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(E2E_HEADERS)
        w.writerows(e2e_rows)

    print(f"CSV files saved to {OUTPUT_DIR}/. Install openpyxl for full XLSX: pip install openpyxl")
    print(f"  UAT: {path_uat}")
    print(f"  SIT: {path_sit}")
    print(f"  E2E: {path_e2e}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Parse behavior specs
    scenarios, spec_count = parse_behavior_specs()

    if not scenarios:
        print("No scenarios found in specs/behavior/REQ-*.md")
        print("Ensure behavior spec files exist with a ## Scenarios table.")
        sys.exit(1)

    # Route to tabs
    uat_scenarios, sit_scenarios, e2e_scenarios = route_scenarios(scenarios)

    # Build rows
    uat_rows = [build_uat_row(s) for s in uat_scenarios]
    sit_rows = [build_sit_row(s) for s in sit_scenarios]
    e2e_rows = [build_e2e_row(s) for s in e2e_scenarios]

    project_name = get_project_name()
    today = date.today().isoformat()

    # Output path
    default_file = f"test-cases-{today}.xlsx"
    output = os.path.join(OUTPUT_DIR, default_file)
    if len(sys.argv) > 2 and sys.argv[1] == "--output":
        output = os.path.join(OUTPUT_DIR, sys.argv[2])

    try:
        import openpyxl  # noqa: F401
        result = generate_xlsx(output, uat_rows, sit_rows, e2e_rows, project_name)
        print(f"Test cases generated: {result}")
        print(f"  UAT: {len(uat_rows)} test cases")
        print(f"  SIT: {len(sit_rows)} test cases")
        print(f"  E2E: {len(e2e_rows)} test cases")
        print(f"  Source: {spec_count} behavior specs")
    except ImportError:
        print("openpyxl not installed. Generating CSV fallback...")
        generate_csv_fallback(uat_rows, sit_rows, e2e_rows)
        print(f"  UAT: {len(uat_rows)} test cases")
        print(f"  SIT: {len(sit_rows)} test cases")
        print(f"  E2E: {len(e2e_rows)} test cases")
        print(f"  Source: {spec_count} behavior specs")
